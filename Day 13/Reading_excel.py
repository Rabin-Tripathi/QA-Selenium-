import openpyxl

# File--> Workbook--> Sheet--> Rows--> Cells

file = "C:/Users/rabin/OneDrive/Desktop/QA.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Test Scenarios"]

rows = sheet.max_row  # Count number of rows in an Excel sheet
cols = sheet.max_column  # Count number of columns in an Excel sheet

# Reading all the rows and columns from Excel sheet
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end='          ')
    print()


