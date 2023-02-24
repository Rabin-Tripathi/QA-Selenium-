import openpyxl

file = "C:/Users/rabin/OneDrive/Desktop/data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

#####  For writing same data
'''
for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "Welcome"

workbook.save(file)
'''

#### For Writing Multiple Data
sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Ram"
sheet.cell(1, 3).value = "Engineer"

sheet.cell(2, 1).value = 11
sheet.cell(2, 2).value = "Hari"
sheet.cell(2, 3).value = "Developer"

sheet.cell(3, 1).value = 10
sheet.cell(3, 2).value = "Rabin"
sheet.cell(3, 3).value = "QA"

workbook.save(file)     # For saving the file after writing
