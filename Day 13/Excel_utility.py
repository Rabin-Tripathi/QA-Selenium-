import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    return sheet.cell(rowNum, colNum).value


def writeData(file, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')

    sheet.cell(rowNum, colNum).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[""]
    redFill = PatternFill(start_color='ff000',
                          end_color='ff000',
                          fill_type='solid')

    sheet.cell(rowNum, colNum).fill = redFill
    workbook.save(file)
